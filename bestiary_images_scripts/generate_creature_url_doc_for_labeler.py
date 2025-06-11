from pymongo import MongoClient
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from urllib.parse import urlparse
import os.path

# Загружаем переменные из .env
load_dotenv()

# Параметры подключения к MongoDB
MONGO_USER = os.getenv("MONGO_INITDB_ROOT_USERNAME")
MONGO_PASS = os.getenv("MONGO_INITDB_ROOT_PASSWORD")
MONGO_HOST = "localhost"
MONGO_PORT = os.getenv("MONGO_PORT", "27017")
MONGO_DB = "bestiary_db"

# Формируем URI и подключаемся
mongo_uri = f"mongodb://{MONGO_USER}:{MONGO_PASS}@{MONGO_HOST}:{MONGO_PORT}/?authSource=admin"
client = MongoClient(mongo_uri)
db = client[MONGO_DB]
collection = db["creatures"]

# Извлекаем все документы
documents = collection.find()

# Списки для данных
url_list = []
filename_list = []

for doc in documents:
    images = doc.get("images", [])
    
    # images[1] гарантированно есть
    url = images[1]
    url_list.append(url)
    
    # images[2] может быть, если есть — вытаскиваем имя файла
    if len(images) >= 3:
        parsed_path = urlparse(images[2]).path
        base = os.path.basename(parsed_path)
        name, _ = os.path.splitext(base)
        filename_list.append(name)
    else:
        filename_list.append("")  # пустая строка для пустой ячейки

# Создаём Excel-файл
wb = Workbook()
ws = wb.active
ws.title = "Creature Images"

# Заголовки
ws.append(["✓", "Image URL", "Filename"])

# Добавляем Data Validation для чекбоксов (TRUE/FALSE)
dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv)

# Заполняем данными
for i, (url, filename) in enumerate(zip(url_list, filename_list), start=2):
    ws[f"A{i}"] = "FALSE"
    ws[f"B{i}"] = url
    ws[f"C{i}"] = filename
    dv.add(ws[f"A{i}"])

# Сохраняем файл
output_filename = "creature_images.xlsx"
wb.save(output_filename)
print(f"Файл сохранён как {output_filename}")
